"""Excel export for Bitácora Monitoreo Regulatorio."""

import io
from datetime import datetime
from typing import List

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src import database as db
from src.utils import setup_logging

logger = setup_logging("exports")

COLUMNS = [
    "No.",
    "Fecha",
    "Circular / Norma",
    "Anexos",
    "Descripción",
    "Fecha Recibido",
    "Fecha de Aplicación",
    "¿Aplica?",
    "Área Relacionada",
    "Encargado Área",
    "Comentario",
    "Estado",
    "Seguimiento",
]

_STATUS_COLORS = {
    "pendiente_revision": "FFF2CC",
    "asignado": "DEEBF7",
    "en_implementacion": "E2EFDA",
    "bloqueado": "FCE4D6",
    "implementado": "C6EFCE",
    "cerrado": "D9D9D9",
}

_RISK_COLORS = {
    "crítico": "FF0000",
    "alto": "FFC000",
    "medio": "FFFF00",
    "bajo": "92D050",
}


def export_bitacora(filters: dict = None) -> bytes:
    items = db.get_all_items(filters or {})
    tracking_map = {r["item_id"]: r for r in db.get_all_tracking()}
    analysis_map = {r["item_id"]: r for r in _get_all_analysis()}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bitácora Regulatoria"

    _write_header(ws)

    for idx, item in enumerate(items, start=1):
        item_id = item["id"]
        tracking = tracking_map.get(item_id, {})
        analysis = analysis_map.get(item_id, {})

        row = [
            idx,
            item.get("publication_date") or item.get("detected_at", "")[:10],
            item.get("title", "")[:255],
            item.get("source_url", ""),
            analysis.get("executive_summary", "")[:500] if analysis else "",
            item.get("detected_at", "")[:10],
            tracking.get("due_date") or analysis.get("max_application_date", ""),
            tracking.get("applies") or analysis.get("applies", ""),
            tracking.get("responsible_area") or analysis.get("suggested_area", ""),
            tracking.get("owner", ""),
            tracking.get("comments", ""),
            tracking.get("progress_status", item.get("status", "")),
            tracking.get("action_plan", ""),
        ]

        ws.append(row)
        excel_row = ws.max_row

        # Color by progress status
        status = tracking.get("progress_status", "")
        fill_hex = _STATUS_COLORS.get(status, "FFFFFF")
        fill = PatternFill("solid", fgColor=fill_hex)
        for col in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=excel_row, column=col)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    _auto_size_columns(ws)
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    logger.info("Excel export generated with %d rows", len(items))
    return buf.getvalue()


def _write_header(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25


def _auto_size_columns(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        adjusted = min(max_len + 4, 60)
        ws.column_dimensions[col_letter].width = adjusted


def _get_all_analysis() -> List[dict]:
    from src.database import get_connection
    with get_connection() as conn:
        rows = conn.execute("SELECT * FROM ai_analysis").fetchall()
        return [dict(r) for r in rows]
